#!/usr/bin/env python3
"""
Excel Analysis Generator
Creates comprehensive USDVND analysis workbook with charts.
"""

import sqlite3
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
import sys

# Add parent for database access
sys.path.insert(0, str(Path(__file__).parent.parent / "scrapers"))

DB_PATH = Path(__file__).parent.parent / "data" / "usdvnd_rates.db"
OUTPUT_PATH = Path(__file__).parent.parent / "analysis" / "USDVND_Analysis.xlsx"


def load_data() -> pd.DataFrame:
    """Load all data from SQLite database."""
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("""
        SELECT * FROM daily_rates
        ORDER BY date
    """, conn)
    conn.close()

    # Convert date column
    df['date'] = pd.to_datetime(df['date'])
    df.set_index('date', inplace=True)

    return df


def compute_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """Compute analysis metrics."""
    # Use best available grey market rate
    df['grey_mid'] = df.apply(
        lambda r: (r['tygiausd_grey_buy'] + r['tygiausd_grey_sell']) / 2
        if pd.notna(r['tygiausd_grey_buy']) and pd.notna(r['tygiausd_grey_sell'])
        else (r['tygiachoden_buy'] + r['tygiachoden_sell']) / 2
        if pd.notna(r['tygiachoden_buy']) and pd.notna(r['tygiachoden_sell'])
        else (r['chogia_buy'] + r['chogia_sell']) / 2
        if pd.notna(r['chogia_buy']) and pd.notna(r['chogia_sell'])
        else None,
        axis=1
    )

    # Official rate (best available)
    df['official_mid'] = df.apply(
        lambda r: r['tygiausd_sbv_central']
        if pd.notna(r['tygiausd_sbv_central'])
        else r['official_usd_vnd']
        if pd.notna(r['official_usd_vnd'])
        else (r['tygiausd_vcb_buy'] + r['tygiausd_vcb_sell']) / 2
        if pd.notna(r['tygiausd_vcb_buy']) and pd.notna(r['tygiausd_vcb_sell'])
        else None,
        axis=1
    )

    # Spread (grey - official)
    df['spread_abs'] = df['grey_mid'] - df['official_mid']
    df['spread_pct'] = (df['spread_abs'] / df['official_mid']) * 100

    # Depreciation metrics (for official rate)
    df['official_pct_change_1d'] = df['official_mid'].pct_change() * 100
    df['official_pct_change_ytd'] = df.groupby(df.index.year)['official_mid'].transform(
        lambda x: (x / x.iloc[0] - 1) * 100
    )
    df['official_pct_change_1y'] = (df['official_mid'] / df['official_mid'].shift(252) - 1) * 100

    # Rolling averages
    df['spread_pct_30d_avg'] = df['spread_pct'].rolling(30).mean()
    df['spread_pct_90d_avg'] = df['spread_pct'].rolling(90).mean()

    # Binance P2P spread
    df['binance_spread'] = df['binance_p2p_sell'] - df['binance_p2p_buy']
    df['binance_spread_pct'] = (df['binance_spread'] / df['binance_p2p_buy']) * 100

    return df


def generate_excel(df: pd.DataFrame):
    """Generate Excel workbook with multiple sheets and charts."""
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.formatting.rule import ColorScaleRule

    wb = Workbook()

    # ===== Sheet 1: Summary Dashboard =====
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Title
    ws_summary['A1'] = "USD/VND Rate Analysis"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A2'] = f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    # Latest rates box
    latest = df.iloc[-1] if len(df) > 0 else None
    if latest is not None:
        ws_summary['A4'] = "LATEST RATES"
        ws_summary['A4'].font = Font(bold=True, size=12)

        metrics = [
            ("Date", latest.name.strftime('%Y-%m-%d') if hasattr(latest.name, 'strftime') else str(latest.name)),
            ("Official (SBV Central)", f"{latest.get('official_mid', 0):,.0f}" if pd.notna(latest.get('official_mid')) else "N/A"),
            ("Grey Market Mid", f"{latest.get('grey_mid', 0):,.0f}" if pd.notna(latest.get('grey_mid')) else "N/A"),
            ("Grey Premium (%)", f"{latest.get('spread_pct', 0):.2f}%" if pd.notna(latest.get('spread_pct')) else "N/A"),
            ("Binance P2P Buy", f"{latest.get('binance_p2p_buy', 0):,.0f}" if pd.notna(latest.get('binance_p2p_buy')) else "N/A"),
            ("Binance P2P Sell", f"{latest.get('binance_p2p_sell', 0):,.0f}" if pd.notna(latest.get('binance_p2p_sell')) else "N/A"),
            ("YTD Depreciation (%)", f"{latest.get('official_pct_change_ytd', 0):.2f}%" if pd.notna(latest.get('official_pct_change_ytd')) else "N/A"),
            ("1Y Depreciation (%)", f"{latest.get('official_pct_change_1y', 0):.2f}%" if pd.notna(latest.get('official_pct_change_1y')) else "N/A"),
        ]

        for i, (label, value) in enumerate(metrics, start=5):
            ws_summary[f'A{i}'] = label
            ws_summary[f'B{i}'] = value
            ws_summary[f'A{i}'].font = Font(bold=True)

        # Spread statistics
        ws_summary['A15'] = "SPREAD STATISTICS"
        ws_summary['A15'].font = Font(bold=True, size=12)

        spread_stats = [
            ("Current Spread (%)", f"{latest.get('spread_pct', 0):.2f}%" if pd.notna(latest.get('spread_pct')) else "N/A"),
            ("30-Day Avg Spread (%)", f"{latest.get('spread_pct_30d_avg', 0):.2f}%" if pd.notna(latest.get('spread_pct_30d_avg')) else "N/A"),
            ("90-Day Avg Spread (%)", f"{latest.get('spread_pct_90d_avg', 0):.2f}%" if pd.notna(latest.get('spread_pct_90d_avg')) else "N/A"),
            ("All-Time Avg Spread (%)", f"{df['spread_pct'].mean():.2f}%" if df['spread_pct'].notna().any() else "N/A"),
            ("All-Time Max Spread (%)", f"{df['spread_pct'].max():.2f}%" if df['spread_pct'].notna().any() else "N/A"),
            ("All-Time Min Spread (%)", f"{df['spread_pct'].min():.2f}%" if df['spread_pct'].notna().any() else "N/A"),
        ]

        for i, (label, value) in enumerate(spread_stats, start=16):
            ws_summary[f'A{i}'] = label
            ws_summary[f'B{i}'] = value
            ws_summary[f'A{i}'].font = Font(bold=True)

    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20

    # ===== Sheet 2: Time Series Data =====
    ws_data = wb.create_sheet("Time Series")

    # Select key columns for export
    cols = ['official_mid', 'grey_mid', 'spread_abs', 'spread_pct',
            'tygiausd_sbv_central', 'tygiausd_grey_buy', 'tygiausd_grey_sell',
            'tygiachoden_buy', 'tygiachoden_sell',
            'binance_p2p_buy', 'binance_p2p_sell',
            'coingecko_usdt_vnd',
            'official_pct_change_ytd', 'official_pct_change_1y']

    export_df = df[[c for c in cols if c in df.columns]].copy()
    export_df = export_df.reset_index()

    # Write headers
    headers = ['Date'] + [c for c in cols if c in df.columns]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

    # Write data
    for row_idx, row in enumerate(export_df.itertuples(), start=2):
        for col_idx, value in enumerate(row[1:], start=1):
            cell = ws_data.cell(row=row_idx, column=col_idx)
            if isinstance(value, pd.Timestamp):
                cell.value = value.strftime('%Y-%m-%d')
            elif pd.isna(value):
                cell.value = None
            else:
                cell.value = value

    # ===== Sheet 3: Charts =====
    ws_charts = wb.create_sheet("Charts")

    # Create rate comparison chart
    if len(export_df) > 10:
        chart1 = LineChart()
        chart1.title = "Official vs Grey Market Rate"
        chart1.style = 10
        chart1.y_axis.title = "VND per USD"
        chart1.x_axis.title = "Date"
        chart1.width = 20
        chart1.height = 10

        # Reference data from Time Series sheet
        data_end = len(export_df) + 1
        official_col = headers.index('official_mid') + 1 if 'official_mid' in headers else None
        grey_col = headers.index('grey_mid') + 1 if 'grey_mid' in headers else None

        if official_col:
            data = Reference(ws_data, min_col=official_col, min_row=1, max_row=data_end)
            chart1.add_data(data, titles_from_data=True)
        if grey_col:
            data = Reference(ws_data, min_col=grey_col, min_row=1, max_row=data_end)
            chart1.add_data(data, titles_from_data=True)

        dates = Reference(ws_data, min_col=1, min_row=2, max_row=data_end)
        chart1.set_categories(dates)

        ws_charts.add_chart(chart1, "A1")

        # Spread chart
        chart2 = LineChart()
        chart2.title = "Grey Market Premium (%)"
        chart2.style = 10
        chart2.y_axis.title = "Spread %"
        chart2.x_axis.title = "Date"
        chart2.width = 20
        chart2.height = 10

        spread_col = headers.index('spread_pct') + 1 if 'spread_pct' in headers else None
        if spread_col:
            data = Reference(ws_data, min_col=spread_col, min_row=1, max_row=data_end)
            chart2.add_data(data, titles_from_data=True)
            chart2.set_categories(dates)
            ws_charts.add_chart(chart2, "A20")

    # ===== Sheet 4: Monthly Summary =====
    ws_monthly = wb.create_sheet("Monthly Summary")

    # Compute monthly averages
    monthly = df.resample('ME').agg({
        'official_mid': 'mean',
        'grey_mid': 'mean',
        'spread_pct': 'mean',
    }).dropna(how='all')

    monthly['month'] = monthly.index.strftime('%Y-%m')

    ws_monthly['A1'] = "Month"
    ws_monthly['B1'] = "Official Rate (Avg)"
    ws_monthly['C1'] = "Grey Rate (Avg)"
    ws_monthly['D1'] = "Spread % (Avg)"

    for col in ['A', 'B', 'C', 'D']:
        ws_monthly[f'{col}1'].font = Font(bold=True)
        ws_monthly[f'{col}1'].fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

    for i, (idx, row) in enumerate(monthly.iterrows(), start=2):
        ws_monthly[f'A{i}'] = row['month']
        ws_monthly[f'B{i}'] = round(row['official_mid'], 0) if pd.notna(row['official_mid']) else None
        ws_monthly[f'C{i}'] = round(row['grey_mid'], 0) if pd.notna(row['grey_mid']) else None
        ws_monthly[f'D{i}'] = round(row['spread_pct'], 2) if pd.notna(row['spread_pct']) else None

    # Adjust column widths
    for col in ['A', 'B', 'C', 'D']:
        ws_monthly.column_dimensions[col].width = 18

    # ===== Sheet 5: Source Comparison =====
    ws_sources = wb.create_sheet("Source Comparison")

    # Latest values from each source
    ws_sources['A1'] = "Source Comparison (Latest Available)"
    ws_sources['A1'].font = Font(bold=True, size=12)

    sources_data = [
        ("Source", "Buy Rate", "Sell Rate", "Last Update"),
        ("tygiausd.org (Grey)", "", "", ""),
        ("tygiachoden.com", "", "", ""),
        ("chogia.vn", "", "", ""),
        ("Binance P2P", "", "", ""),
        ("CoinGecko (USDT)", "", "", ""),
        ("SBV Central", "", "", ""),
        ("Vietcombank", "", "", ""),
    ]

    for i, row in enumerate(sources_data, start=3):
        for j, val in enumerate(row):
            ws_sources.cell(row=i, column=j+1, value=val)
            if i == 3:
                ws_sources.cell(row=i, column=j+1).font = Font(bold=True)

    # Fill in latest values
    if latest is not None:
        row_map = {
            4: ('tygiausd_grey_buy', 'tygiausd_grey_sell'),
            5: ('tygiachoden_buy', 'tygiachoden_sell'),
            6: ('chogia_buy', 'chogia_sell'),
            7: ('binance_p2p_buy', 'binance_p2p_sell'),
            8: ('coingecko_usdt_vnd', 'coingecko_usdt_vnd'),
            9: ('tygiausd_sbv_central', 'tygiausd_sbv_central'),
            10: ('tygiausd_vcb_buy', 'tygiausd_vcb_sell'),
        }

        for row_num, (buy_col, sell_col) in row_map.items():
            buy_val = latest.get(buy_col)
            sell_val = latest.get(sell_col)
            ws_sources.cell(row=row_num, column=2, value=f"{buy_val:,.0f}" if pd.notna(buy_val) else "N/A")
            ws_sources.cell(row=row_num, column=3, value=f"{sell_val:,.0f}" if pd.notna(sell_val) else "N/A")

    # Save
    wb.save(OUTPUT_PATH)
    print(f"Excel file saved to: {OUTPUT_PATH}")


def main():
    """Main function to generate analysis."""
    print("Loading data from database...")
    df = load_data()
    print(f"Loaded {len(df)} records")

    if len(df) == 0:
        print("No data available. Run backfill first.")
        return

    print("Computing metrics...")
    df = compute_metrics(df)

    print("Generating Excel workbook...")
    generate_excel(df)

    print("Done!")


if __name__ == "__main__":
    main()
